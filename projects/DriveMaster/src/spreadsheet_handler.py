import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import Rule, FormulaRule

def write_report_to_csv(report_data, filename):
    """Writes report data to a CSV file. Returns True on success, False on failure."""
    if not report_data:
        logging.warning("No data to write to CSV report.")
        return True
    try:
        df = pd.DataFrame(report_data)
        df.to_csv(filename, index=False, encoding='utf-8')
        return True
    except Exception as e:
        logging.error(f"Failed to write backup CSV to {filename}: {e}")
        return False

def save_audit_log(audit_data, filename):
    """Saves the audit trail to a CSV log file. Returns True on success, False on failure."""
    if not isinstance(audit_data, list):
        logging.error(f"Invalid data type for audit log: expected list, got {type(audit_data)}")
        return False
    if not audit_data:
        logging.warning("No audit data to write to log file.")
        return True
    try:
        df = pd.DataFrame(audit_data)
        if df.empty:
            logging.warning("Audit DataFrame is empty after conversion.")
            return True
        log_column_order = ['Timestamp', 'Root Folder ID', 'Full Path', 'Item Name', 'Item ID', 'Action_Command', 'Status', 'Details', 'Original_Principal_Type', 'Original_Email_Address', 'Original_Role', 'New_Principal_Type', 'New_Email_Address', 'New_Role']
        df = df.reindex(columns=log_column_order)
        df.to_csv(filename, index=False, encoding='utf-8')
        logging.info(f"Audit log successfully written to {filename}")
        return True
    except Exception as e:
        logging.error(f"Failed to write audit log to {filename}: {e}")
        return False

def add_dropdowns_to_sheet(filename, domain_filter=None):
    """
    Adds dropdowns, formatting, and auto-filter to an Excel sheet.
    Optionally applies a filter to the 'Email Address' column.
    """
    try:
        wb = load_workbook(filename)
        ws = wb.active
        
        ws.auto_filter.ref = ws.dimensions
        
        # --- MODIFIED: Apply domain filter if provided ---
        if domain_filter:
            email_col_index = -1
            # Find the 'Email Address' column index (1-based)
            for i, cell in enumerate(ws[1]):
                if cell.value == 'Email Address':
                    email_col_index = i + 1
                    break
            
            if email_col_index != -1:
                logging.info(f"Applying Excel filter for domain: *@{domain_filter}")
                # Add a filter to the column by its index
                ws.auto_filter.add_filter_column(email_col_index - 1, [f"*@{domain_filter}"])

        dv_set_restrict = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
        dv_action = DataValidation(type="list", formula1='"MODIFY,REMOVE,ADD"', allow_blank=True)
        dv_role = DataValidation(type="list", formula1='"Viewer,Commenter,Editor"', allow_blank=True)
        dv_principal = DataValidation(type="list", formula1='"user,group,domain"', allow_blank=True)
        
        # Columns shifted due to new 'Mime Type' column
        dv_action.add('L2:L1048576')       # Action_Type
        dv_role.add('M2:M1048576')       # New_Role
        dv_principal.add('N2:N1048576')  # Type of account (for ADD)
        dv_set_restrict.add('P2:P1048576') # SET Download Restriction
        
        ws.add_data_validation(dv_action)
        ws.add_data_validation(dv_role)
        ws.add_data_validation(dv_principal)
        ws.add_data_validation(dv_set_restrict)
        
        fill_add = PatternFill(start_color="FFD8E9BB", end_color="FFD8E9BB", fill_type="solid")
        fill_remove = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
        fill_modify = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")
        
        full_range = 'A2:P1048576' # Range extended
        ws.conditional_formatting.add(full_range, FormulaRule(formula=['=$L2="ADD"'], fill=fill_add))
        ws.conditional_formatting.add(full_range, FormulaRule(formula=['=$L2="REMOVE"'], fill=fill_remove))
        ws.conditional_formatting.add(full_range, FormulaRule(formula=['=$L2="MODIFY"'], fill=fill_modify))
        
        wb.save(filename)
        logging.info(f"Successfully added auto-filter, dropdowns, and formatting to {filename}")
    except Exception as e:
        logging.error(f"Could not add dropdowns or formatting to {filename}. Reason: {e}")
        # Do not re-raise; allow the process to continue if this non-critical step fails.

def write_report_to_excel(report_data, filename, domain_filter=None):
    """
    Writes the report data to an Excel file.
    Optionally passes a domain filter to the sheet formatter.
    """
    if not isinstance(report_data, list):
        logging.error(f"Invalid data type for Excel report: expected list, got {type(report_data)}")
        return False
    if not report_data:
        logging.warning("No data to write to Excel report.")
        return True
    
    try:
        df = pd.DataFrame(report_data)
        
        action_columns = ['Action_Type', 'New_Role', 'Type of account (for ADD)', 'Email/Domain (for ADD)', 'SET Download Restriction']
        for col in action_columns: df[col] = ''
        
        column_order = ['Full Path', 'Item Name', 'Item ID', 'Mime Type', 'Role', 'Principal Type', 'Email Address', 'Owner', 'Google Drive URL', 'Root Folder ID', 'Current Download Restriction'] + action_columns
        
        df = df.reindex(columns=column_order)
        df.to_excel(filename, index=False, engine='openpyxl')
        
        # --- MODIFIED: Pass the domain filter to the formatter ---
        add_dropdowns_to_sheet(filename, domain_filter)
        return True
    except Exception as e:
        logging.error(f"Failed to write report to Excel file {filename}: {e}")
        # Re-raise the exception for the GUI to catch and handle (e.g., for PermissionError)
        raise

